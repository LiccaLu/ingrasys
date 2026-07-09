import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urljoin, quote
import re
from io import BytesIO
from datetime import date
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

st.set_page_config(
    page_title="勞動法規行政規則整理工具",
    page_icon="⚖️",
    layout="wide"
)

st.markdown("""
<style>
.block-container {
    padding-top: 2rem;
    max-width: 1200px;
}

h1 {
    font-size: 42px !important;
    font-weight: 800 !important;
}

.stButton > button {
    background: #1f7a3f;
    color: white;
    border-radius: 10px;
    padding: 0.6rem 1.4rem;
    border: none;
    font-weight: 700;
}

.stButton > button:hover {
    background: #166534;
    color: white;
}

div[data-testid="stDateInput"] input {
    border-radius: 10px;
}

div[data-testid="stAlert"] {
    border-radius: 14px;
}

[data-testid="stDataFrame"] {
    border-radius: 14px;
    overflow: hidden;
    border: 1px solid #e5e7eb;
}

.stDownloadButton > button {
    background: #2563eb;
    color: white;
    border-radius: 10px;
    padding: 0.6rem 1.4rem;
    font-weight: 700;
    border: none;
}
</style>
""", unsafe_allow_html=True)


BASE_URL = "https://laws.mol.gov.tw"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
    "Referer": "https://laws.mol.gov.tw/",
    "Connection": "close",
}

connection_failed = False


def get_html(url, timeout=25):
    global connection_failed

    try:
        r = requests.get(
            url,
            headers=HEADERS,
            verify=False,
            timeout=timeout
        )
        r.raise_for_status()
        r.encoding = "utf-8"
        return BeautifulSoup(r.text, "html.parser")

    except requests.exceptions.Timeout:
        connection_failed = True
        st.warning(f"連線逾時，略過：{url}")
        return None

    except requests.exceptions.RequestException as e:
        connection_failed = True
        st.warning(f"無法連線，略過：{url}")
        st.caption(repr(e))
        return None


def roc_to_date(roc_text):
    y, m, d = roc_text.split(".")
    return date(int(y) + 1911, int(m), int(d))


def clean_title(title):
    title = title.replace("勞動部令：", "")
    title = title.replace("勞動部公告：", "")
    title = title.strip()

    effective_date = ""
    match = re.search(r"自(.+?)生效", title)

    if match:
        effective_date = match.group(1)
        title = title.split("，自")[0]

    return title, effective_date


def scrape_index(start_date, end_date):
    records = []
    page = 1

    while True:
        if page == 1:
            url = BASE_URL + "/index.aspx"
        else:
            url = BASE_URL + f"/index.aspx?page={page}"

        soup = get_html(url)

        if soup is None:
            st.error("連不上勞動部網站。這不是沒有資料，是目前部署環境無法連到 laws.mol.gov.tw。")
            st.stop()

        table = soup.find("table", class_="table-list news-table")

        if table is None:
            break

        should_stop = False

        for row in table.find_all("tr")[1:]:
            cols = row.find_all("td")

            if len(cols) < 3:
                continue

            date_text = cols[0].get_text(strip=True)

            try:
                publish_date = roc_to_date(date_text)
            except Exception:
                continue

            if publish_date < start_date:
                should_stop = True
                break

            if publish_date > end_date:
                continue

            category = cols[1].get_text(strip=True)

            if category != "行政規則":
                continue

            raw_title = cols[2].get_text(strip=True)
            title, effective_date = clean_title(raw_title)

            a = cols[2].find("a")
            notice_url = urljoin(BASE_URL, a.get("href", "")) if a else ""

            records.append({
                "公發布日": date_text,
                "類別": category,
                "訊息摘要": title,
                "生效日期": effective_date,
                "公告連結": notice_url
            })

        if should_stop:
            break

        page += 1

    return pd.DataFrame(records)


def get_law_name(detail_soup):
    if detail_soup is None:
        return ""

    lines = detail_soup.get_text("\n", strip=True).split("\n")

    for i, line in enumerate(lines):
        if "法規名稱" in line and i + 1 < len(lines):
            return lines[i + 1].strip()

    return ""


def get_gazette_url(detail_url, detail_soup):
    if detail_soup is None:
        return ""

    for a in detail_soup.find_all("a"):
        text = a.get_text(strip=True)
        href = a.get("href", "")

        if "行政院公報" in text and href:
            return urljoin(detail_url, href.replace("&amp;", "&"))

    return ""


def get_text_version_url(gazette_url):
    if not gazette_url:
        return ""

    soup = get_html(gazette_url)

    if soup is None:
        return ""

    for tag in soup.find_all(["a", "iframe"]):
        text = tag.get_text(strip=True)
        href = tag.get("href") or tag.get("src") or ""

        if not href:
            continue

        if (
            "網頁文字版" in text
            or "文字版" in text
            or "eguploadpubWrapper" in href
            or "eguploadpub" in href
        ):
            return urljoin(gazette_url, href.replace("&amp;", "&"))

    return ""


def search_history(law_name):
    if not law_name:
        return "", 0

    search_url = BASE_URL + "/results.aspx?searchmode=global&keyword=" + quote(law_name)

    soup = get_html(search_url)

    if soup is None:
        return "", 0

    dates = []

    for row in soup.find_all("tr"):
        cols = row.find_all("td")

        if len(cols) < 4:
            continue

        category = cols[1].get_text(strip=True)
        date_text = cols[3].get_text(strip=True)

        if category != "最新動態":
            dates.append(date_text)

    if dates:
        return "、".join(dates), len(dates)

    return "", 0


def enrich_one(row):
    notice_url = row["公告連結"]

    law_name = ""
    history_dates = ""
    history_count = 0
    gazette_url = ""
    text_url = ""

    try:
        detail_soup = get_html(notice_url)

        if detail_soup is not None:
            law_name = get_law_name(detail_soup)
            gazette_url = get_gazette_url(notice_url, detail_soup)
            text_url = get_text_version_url(gazette_url)
            history_dates, history_count = search_history(law_name)

    except Exception:
        pass

    return {
        "法規名稱": law_name,
        "歷次修改日期": history_dates,
        "歷史筆數": history_count,
        "行政院公報連結": gazette_url,
        "網頁文字版連結": text_url
    }


def scrape_laws(start_date, end_date):
    df = scrape_index(start_date, end_date)

    if df.empty:
        return df

    rows = [row for _, row in df.iterrows()]

    with ThreadPoolExecutor(max_workers=1) as executor:
        enriched = list(executor.map(enrich_one, rows))

    enrich_df = pd.DataFrame(enriched)
    df = pd.concat([df.reset_index(drop=True), enrich_df], axis=1)

    return df


def to_excel(df):
    output = BytesIO()

    export_df = df.copy()

    columns_order = [
        "公發布日",
        "類別",
        "法規名稱",
        "訊息摘要",
        "生效日期",
        "歷次修改日期",
        "歷史筆數",
        "公告連結",
        "行政院公報連結",
        "網頁文字版連結"
    ]

    export_df = export_df[[c for c in columns_order if c in export_df.columns]]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(
            writer,
            index=False,
            sheet_name="行政規則整理",
            startrow=4
        )

        wb = writer.book
        ws = writer.sheets["行政規則整理"]

        dark_blue = "1F4E79"
        light_gray = "EAF0F6"
        white = "FFFFFF"
        border_gray = "BFBFBF"

        thin_border = Border(
            left=Side(style="thin", color=border_gray),
            right=Side(style="thin", color=border_gray),
            top=Side(style="thin", color=border_gray),
            bottom=Side(style="thin", color=border_gray)
        )

        max_col = len(export_df.columns)

        # 報表大標題
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws["A1"] = "勞動部行政規則整理報告"
        ws["A1"].font = Font(size=18, bold=True, color=dark_blue)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        # 查詢期間
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        ws["A2"] = f"查詢期間：{start_date.strftime('%Y/%m/%d')} ~ {end_date.strftime('%Y/%m/%d')}｜共 {len(export_df)} 筆"
        ws["A2"].font = Font(size=11, bold=True, color="404040")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # 產生時間
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        ws["A3"] = f"產生時間：{datetime.now().strftime('%Y/%m/%d %H:%M')}"
        ws["A3"].font = Font(size=10, color="666666")
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        header_row = 5

        # 標題列
        for cell in ws[header_row]:
            cell.fill = PatternFill("solid", fgColor=dark_blue)
            cell.font = Font(color=white, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 內容列
        for row in range(header_row + 1, ws.max_row + 1):
            fill_color = light_gray if row % 2 == 0 else white

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = thin_border
                cell.font = Font(size=10)
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True
                )

        # 超連結改成文字
        link_columns = {
            "公告連結": "開啟公告",
            "行政院公報連結": "開啟公報",
            "網頁文字版連結": "開啟文字版"
        }

        headers = [cell.value for cell in ws[header_row]]

        for col_name, display_text in link_columns.items():
            if col_name in headers:
                col_idx = headers.index(col_name) + 1

                for row in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    url = cell.value

                    if url:
                        cell.value = display_text
                        cell.hyperlink = url
                        cell.style = "Hyperlink"
                        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 欄寬
        column_widths = {
            "公發布日": 12,
            "類別": 12,
            "法規名稱": 38,
            "訊息摘要": 55,
            "生效日期": 18,
            "歷次修改日期": 28,
            "歷史筆數": 10,
            "公告連結": 14,
            "行政院公報連結": 14,
            "網頁文字版連結": 14
        }

        for idx, col_name in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = column_widths.get(col_name, 18)

        # 列高
        ws.row_dimensions[1].height = 32
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 20
        ws.row_dimensions[header_row].height = 35

        for row in range(header_row + 1, ws.max_row + 1):
            ws.row_dimensions[row].height = 70

        # 凍結標題列
        ws.freeze_panes = "A6"

        # 篩選
        ws.auto_filter.ref = f"A5:{get_column_letter(ws.max_column)}{ws.max_row}"

        # 隱藏格線
        ws.sheet_view.showGridLines = False

        # 橫向列印
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    return output.getvalue()


st.title("⚖️ 勞動法規行政規則整理工具")
st.write("選擇日期區間後，系統會即時爬取勞動部最新動態，整理行政規則，並產生 Excel 下載檔。")

start_date = st.date_input("開始日期", value=date(2026, 6, 1))
end_date = st.date_input("結束日期", value=date(2026, 7, 7))

if start_date > end_date:
    st.error("開始日期不能晚於結束日期。")
    st.stop()


if st.button("開始整理"):
    connection_failed = False

    with st.spinner("正在爬取資料，請稍候..."):
        df = scrape_laws(start_date, end_date)

    if df.empty:
        st.warning("成功連上網站，但這個日期區間沒有找到行政規則。")
    else:
        st.success(f"整理完成，共 {len(df)} 筆行政規則")

        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "公告連結": st.column_config.LinkColumn(
                    "公告連結",
                    display_text="開啟公告"
                ),
                "行政院公報連結": st.column_config.LinkColumn(
                    "行政院公報連結",
                    display_text="開啟公報"
                ),
                "網頁文字版連結": st.column_config.LinkColumn(
                    "網頁文字版連結",
                    display_text="開啟文字版"
                ),
            }
        )

        st.download_button(
            label="下載 Excel 檔案",
            data=to_excel(df),
            file_name="行政規則公報整理.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
